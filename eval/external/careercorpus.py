"""CareerCorpus: correlate the live scorer against dual expert annotations.

    python careercorpus.py --file data/careercorpus.xlsx

What this can validate
  Whether the scorer's *quality* ranking agrees with human experts.

What it CANNOT validate
  Job matching. CareerCorpus ships no job descriptions, so every resume is scored
  with job_description=None. The semantic component (20% of the total weight) then
  returns its neutral constant for all 302 rows and contributes zero variance. The
  correlation reported here is therefore a verdict on the other 80% only.

Read the verdict against the inter-annotator ceiling printed at the top, never
against 1.0. On four of six categories the two human experts agree with each other
less than 0.70; a scorer cannot meaningfully exceed the humans who labelled it.
"""
from __future__ import annotations

import argparse
import re
import statistics
import sys
from collections import defaultdict
from pathlib import Path

from _harness import correlations, distribution_report, histogram, sbert_mode, score_text

# Published inter-annotator agreement, from the CareerCorpus paper. This is the
# ceiling: the experts themselves disagreed by this much.
PUBLISHED_IAA = {
    "Apparel": 0.89,
    "Finance": 0.68,
    "Research Assistant": 0.67,
    "Teacher": 0.56,
    "Banking": 0.38,
    "Accountant": 0.35,
}


def load_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required to read the .xlsx:\n"
                 "  backend/.venv/Scripts/pip install openpyxl")

    if not path.exists():
        sys.exit(f"Not found: {path}\nDownload it first — see external/README.md")

    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    return [dict(zip(header, row)) for row in rows]


def detect_columns(records: list[dict]) -> tuple[str, list[str], str | None]:
    """Find the resume text, the two annotator score columns and the category.

    The published column names are not guaranteed stable across dataset versions,
    so they are inferred and then printed for confirmation rather than hardcoded.
    Override with --text-col / --score-cols / --category-col.
    """
    header = list(records[0].keys())
    sample = records[: min(40, len(records))]

    def avg_len(col: str) -> float:
        return statistics.fmean(len(str(r.get(col) or "")) for r in sample)

    def numeric_share(col: str) -> float:
        vals = [r.get(col) for r in sample]
        return sum(1 for v in vals if isinstance(v, (int, float))) / max(1, len(vals))

    def looks_like_id(col: str) -> bool:
        """A row counter is numeric and near-unique — never an annotation score."""
        if re.fullmatch(r"(unnamed:?\s*\d*|s\.?\s*no\.?|sr\.?\s*no\.?|id|index|no)", col.strip(), re.I):
            return True
        vals = [r.get(col) for r in sample if isinstance(r.get(col), (int, float))]
        return len(vals) > 5 and len(set(vals)) / len(vals) > 0.9

    text_col = max(header, key=avg_len)

    numeric = [c for c in header
               if c != text_col and numeric_share(c) > 0.8 and avg_len(c) < 12
               and not looks_like_id(c)]
    # A name hint beats the numeric heuristic: the two expert columns are the ones
    # actually called something like "Annotator 1" / "Expert Score", and picking an
    # unrelated numeric column here silently correlates the scorer against noise.
    named = [c for c in numeric
             if re.search(r"annot|score|rating|rate|expert|grade|eval|label", c, re.I)]
    score_cols = (named or numeric)[:2]

    category_col = None
    for col in header:
        if col in (text_col, *score_cols) or numeric_share(col) > 0.8:
            continue
        values = {str(r.get(col)) for r in sample}
        if 2 <= len(values) <= 12 and avg_len(col) < 40:
            category_col = col
            break

    return text_col, score_cols, category_col


def verdict(observed: float, ceiling: float | None) -> str:
    if observed != observed:  # NaN
        return "undefined (no variance in one series)"
    if ceiling is None:
        return "no published ceiling for this category"
    if observed >= ceiling:
        return f"{observed:.2f} vs human ceiling {ceiling:.2f} — AT OR ABOVE expert agreement"
    if observed >= ceiling * 0.75:
        return f"{observed:.2f} vs human ceiling {ceiling:.2f} — within expert disagreement range"
    if observed >= ceiling * 0.4:
        return f"{observed:.2f} vs human ceiling {ceiling:.2f} — below experts, same direction"
    if observed > 0.1:
        return f"{observed:.2f} vs human ceiling {ceiling:.2f} — WEAK, far under the ceiling"
    return f"{observed:.2f} vs human ceiling {ceiling:.2f} — NO USABLE SIGNAL"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", type=Path, default=Path("data/careercorpus.xlsx"))
    ap.add_argument("--text-col")
    ap.add_argument("--score-cols", nargs=2)
    ap.add_argument("--category-col")
    ap.add_argument("--limit", type=int, help="score only the first N rows (smoke test)")
    args = ap.parse_args()

    records = load_xlsx(args.file)
    text_col, score_cols, category_col = detect_columns(records)
    text_col = args.text_col or text_col
    score_cols = args.score_cols or score_cols
    category_col = args.category_col or category_col

    if len(score_cols) < 2:
        sys.exit(f"Could not find two annotator score columns. Columns present:\n"
                 f"  {list(records[0].keys())}\nPass them with --score-cols A B")

    print("=" * 78)
    print("CAREERCORPUS — live scorer vs dual expert annotation".center(78))
    print("=" * 78)
    print(f"file           {args.file}")
    print(f"rows           {len(records)}")
    print(f"text column    {text_col!r}")
    print(f"score columns  {score_cols[0]!r}, {score_cols[1]!r}")
    print(f"category col   {category_col!r}")
    print(f"semantic path  {sbert_mode()}")

    # ---------------------------------------------------------------- ceiling
    print("\n" + "=" * 78)
    print("INTER-ANNOTATOR AGREEMENT — THIS IS THE CEILING".center(78))
    print("=" * 78)
    print("Published agreement between the two expert annotators. The scorer is")
    print("judged against these numbers, NOT against 1.0.\n")
    for name, value in sorted(PUBLISHED_IAA.items(), key=lambda kv: -kv[1]):
        bar = "█" * int(value * 40)
        print(f"  {name:<20} {value:.2f}  {bar}")
    print(f"\n  mean ceiling across categories: "
          f"{statistics.fmean(PUBLISHED_IAA.values()):.2f}")
    print("  On Banking (0.38) and Accountant (0.35) the experts barely agreed at")
    print("  all — a scorer matching 0.35 there has matched human performance.")

    # measured agreement in this copy of the data
    a_all = [r[score_cols[0]] for r in records if isinstance(r.get(score_cols[0]), (int, float))]
    b_all = [r[score_cols[1]] for r in records if isinstance(r.get(score_cols[1]), (int, float))]
    if len(a_all) == len(b_all) and len(a_all) > 2:
        rho, r = correlations(a_all, b_all)
        print(f"\n  measured in this file: spearman {rho:.2f}, pearson {r:.2f} "
              f"(n={len(a_all)})")

    # ---------------------------------------------------------------- scoring
    print("\n" + "=" * 78)
    print("SCORING".center(78))
    print("=" * 78)
    print("!! No job descriptions exist in this dataset, so job_description=None.")
    print("!! The semantic component (20% of total weight) returns its neutral")
    print("!! constant for every row and adds zero variance. What follows measures")
    print("!! the remaining 80%: skills, experience, projects, education, formatting.\n")

    rows = records[: args.limit] if args.limit else records
    scored: list[tuple[str, float, int]] = []
    skipped = 0
    for i, rec in enumerate(rows, 1):
        text = str(rec.get(text_col) or "").strip()
        a, b = rec.get(score_cols[0]), rec.get(score_cols[1])
        if not text or not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
            skipped += 1
            continue
        category = str(rec.get(category_col) or "Uncategorised").strip()
        scored.append((category, (a + b) / 2, score_text(text)["total"]))
        if i % 25 == 0:
            print(f"  scored {i}/{len(rows)}…")

    if skipped:
        print(f"  skipped {skipped} rows with missing text or annotations")
    if len(scored) < 3:
        sys.exit("Too few usable rows to correlate.")

    # ---------------------------------------------------------------- results
    print("\n" + "=" * 78)
    print("CORRELATION vs MEAN EXPERT SCORE".center(78))
    print("=" * 78)
    print(f"{'category':<22} {'n':>4} {'spearman':>9} {'pearson':>8}   verdict")
    print("-" * 78)

    by_category: dict[str, list[tuple[float, int]]] = defaultdict(list)
    for category, expert, ours in scored:
        by_category[category].append((expert, ours))

    for category, pairs in sorted(by_category.items()):
        if len(pairs) < 3:
            print(f"{category:<22} {len(pairs):>4}   too few rows to correlate")
            continue
        rho, r = correlations([p[0] for p in pairs], [p[1] for p in pairs])
        ceiling = PUBLISHED_IAA.get(category)
        print(f"{category:<22} {len(pairs):>4} {rho:>9.2f} {r:>8.2f}   "
              f"{verdict(rho, ceiling)}")

    rho_all, r_all = correlations([s[1] for s in scored], [s[2] for s in scored])
    mean_ceiling = statistics.fmean(PUBLISHED_IAA.values())
    print("-" * 78)
    print(f"{'OVERALL':<22} {len(scored):>4} {rho_all:>9.2f} {r_all:>8.2f}   "
          f"{verdict(rho_all, mean_ceiling)}")

    # ---------------------------------------------------------------- spread
    totals = [s[2] for s in scored]
    print("\n" + "=" * 78)
    print("SCORE DISTRIBUTION".center(78))
    print("=" * 78)
    print(distribution_report(totals))
    print()
    print(histogram(totals))
    print("\nNote: CareerCorpus resumes are professionally crafted (LiveCareer), so the")
    print("true quality range is narrow by construction. A compressed output band is")
    print("partly the dataset and partly the scorer — this test cannot separate them.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
